#!/usr/bin/env python3
"""
Script that reads SAP customer data from `VCUST.XLSX`, removes rows for SAP IDs already
present in the GoDam database, and writes a reduced workbook containing only the header,
`Customer`, and `Name 1` columns for the new records.

Usage:
  python scripts/prepare_vcust.py /path/to/VCUST.XLSX --output ./out/new.csv \
      [--db-url postgresql://user:pass@host:port/dbname]

Set `DATABASE_URL` if you prefer not to pass `--db-url`. Install dependencies with:
  pip install openpyxl psycopg2-binary
"""

from __future__ import annotations

import argparse
import os
import sys
from typing import Iterable, Sequence

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover
    raise SystemExit(
        "openpyxl is required to run this script. Install it via "
        "`pip install openpyxl psycopg2-binary`."
    ) from exc


def normalize_header(value: str | None) -> str:
    if value is None:
        return ""
    return value.strip().lower()


def find_column_index(
    headers: Sequence[str], candidates: Iterable[str]
) -> int | None:
    candidate_set = {normalize_header(candidate) for candidate in candidates}
    for index, header in enumerate(headers):
        if normalize_header(header) in candidate_set:
            return index
    return None


def read_existing_sap_ids(db_url: str) -> set[str]:
    try:
        import psycopg2  # type: ignore
    except ImportError as exc:  # pragma: no cover
        raise SystemExit(
            "psycopg2-binary is required to query the database. "
            "Install it via `pip install psycopg2-binary`."
        ) from exc

    sap_ids: set[str] = set()
    with psycopg2.connect(db_url) as connection:
        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT sap_customer_id FROM customers WHERE sap_customer_id IS NOT NULL"
            )
            for row in cursor:
                raw_value = row[0]
                if raw_value:
                    sap_ids.add(str(raw_value).strip().lower())
    return sap_ids


def prepare_rows(
    workbook: openpyxl.Workbook,
    customer_column: int,
    name_column: int,
    existing_sap_ids: set[str],
) -> tuple[list[tuple[str, str]], int, int]:
    sheet = workbook.active
    new_rows: list[tuple[str, str]] = []
    seen: set[str] = set()
    duplicates_in_file = 0
    duplicates_in_db = 0
    for row in sheet.iter_rows(min_row=2):
        raw_customer = row[customer_column].value
        if raw_customer is None:
            continue
        customer = str(raw_customer).strip()
        if not customer:
            continue
        normalized = customer.lower()
        if normalized in seen:
            duplicates_in_file += 1
            continue
        seen.add(normalized)
        if normalized in existing_sap_ids:
            duplicates_in_db += 1
            continue
        name_value = row[name_column].value if name_column is not None else ""
        name_text = str(name_value).strip() if name_value else ""
        new_rows.append((customer, name_text))
    return new_rows, duplicates_in_file, duplicates_in_db


def write_output(output_path: str, rows: list[tuple[str, str]]) -> None:
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "NewCustomers"
    out_ws.append(["Customer", "Name 1"])
    for customer, name in rows:
        out_ws.append([customer, name])
    out_ws.column_dimensions["A"].width = 30
    out_ws.column_dimensions["B"].width = 30
    out_wb.save(output_path)


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Filter VCUST.XLSX for rows with SAP IDs that are not in the database."
    )
    parser.add_argument(
        "input",
        help="Path to the VCUST Excel file.",
    )
    parser.add_argument(
        "--output",
        default="vcust-new.xlsx",
        help="Output Excel file that will contain only the new rows.",
    )
    parser.add_argument(
        "--db-url",
        default=os.environ.get("DATABASE_URL"),
        help="Postgres connection string; defaults to DATABASE_URL env var.",
    )
    parser.add_argument(
        "--customer-column",
        default="Customer",
        help="Header name for the SAP customer ID column.",
    )
    parser.add_argument(
        "--name-column",
        default="Name 1",
        help="Header name for the human-readable name column.",
    )
    parser.add_argument(
        "--skip-db-check",
        action="store_true",
        help="Only deduplicate within the workbook; do not consult the database.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_arguments()

    if not os.path.isfile(args.input):
        raise SystemExit(f"Input file '{args.input}' does not exist.")

    workbook = openpyxl.load_workbook(args.input, read_only=True)
    sheet = workbook.active
    headers = [cell.value if cell.value is not None else "" for cell in sheet[1]]

    customer_idx = find_column_index(
        headers, (args.customer_column, "customer")
    )
    if customer_idx is None:
        raise SystemExit(f"Unable to find a '{args.customer_column}' column in the workbook.")

    name_idx = find_column_index(headers, (args.name_column,)) or customer_idx

    existing_ids = set()
    if not args.skip_db_check:
        if not args.db_url:
            raise SystemExit(
                "Database URL is required to check for existing SAP IDs unless --skip-db-check is set."
            )
        print("Querying database for existing SAP customer IDs...")
        existing_ids = read_existing_sap_ids(args.db_url)
        print(f"Found {len(existing_ids)} SAP IDs in the database.")

    workbook = openpyxl.load_workbook(args.input, read_only=True)
    rows, dup_file, dup_db = prepare_rows(workbook, customer_idx, name_idx, existing_ids)
    if not rows:
        print("No new rows found.")
        return

    output_dir = os.path.dirname(args.output)
    if output_dir and not os.path.isdir(output_dir):
        os.makedirs(output_dir, exist_ok=True)
    write_output(args.output, rows)

    print(f"Filtered {len(rows)} rows out of duplicates (file: {dup_file}, DB: {dup_db}).")
    print(f"Saved new rows to {args.output}")


if __name__ == "__main__":
    main()
